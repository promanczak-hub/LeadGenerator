"""
Converts non-PDF document formats to text for Gemini API compatibility.

Gemini API supports PDF natively but not XLSX/XLS. This module
converts spreadsheet files into markdown-formatted text that can
be sent to Gemini as a text part.

Handles merged cells by propagating values across merged ranges.
"""

import io
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


_XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}


def is_spreadsheet(mime_type: str) -> bool:
    """Check if the MIME type is a spreadsheet format."""
    return mime_type in _XLSX_MIME_TYPES


def _unmerge_and_fill(ws: Worksheet) -> None:
    """
    Unmerge all merged cell ranges and fill each cell
    with the value from the top-left corner of its merge range.

    This ensures no data is lost during conversion to text.
    Modifies the worksheet in-place.
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        # Get the value from the top-left cell
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        top_left_value = ws.cell(row=min_row, column=min_col).value

        # Unmerge first, then fill
        ws.unmerge_cells(str(merged_range))

        # Fill all cells in the former range with the top-left value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def xlsx_bytes_to_markdown(file_bytes: bytes) -> str:
    """
    Convert XLSX file bytes into a markdown-formatted text representation.

    Each sheet becomes a section with a markdown table.
    Merged cells are unmerged and filled before conversion.
    Empty rows are skipped to keep output concise.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge and fill merged cells
        _unmerge_and_fill(ws)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Determine the max number of columns across all rows
        max_cols = max(len(row) for row in rows)

        # Filter out completely empty rows
        non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]

        if not non_empty_rows:
            continue

        section_lines = [f"## Arkusz: {sheet_name}\n"]

        for row_idx, row in enumerate(non_empty_rows):
            # Pad row to max_cols for consistent table width
            padded = list(row) + [None] * (max_cols - len(row))
            cells = [str(cell) if cell is not None else "" for cell in padded]
            line = "| " + " | ".join(cells) + " |"
            section_lines.append(line)

            # Add header separator after first row
            if row_idx == 0:
                separator = "| " + " | ".join(["---"] * len(cells)) + " |"
                section_lines.append(separator)

        sections.append("\n".join(section_lines))

    wb.close()

    if not sections:
        return "(Pusty plik XLSX — brak danych)"

    header = "# Zawartość dokumentu XLSX\n\n"
    return header + "\n\n".join(sections)


def convert_to_gemini_input(
    file_bytes: bytes, mime_type: str
) -> tuple[str | bytes, str]:
    """
    Convert file bytes to a format Gemini can process.

    Returns (data, effective_mime_type):
      - For PDFs: returns original bytes + original mime_type
      - For XLSX/XLS: returns markdown text + "text/plain"
    """
    if is_spreadsheet(mime_type):
        text = xlsx_bytes_to_markdown(file_bytes)
        return text, "text/plain"

    return file_bytes, mime_type
