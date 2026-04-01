"""Excel → universal_features importer.

Reads the fleet features Excel file and populates:
- reverse_search.universal_features
- reverse_search.universal_feature_aliases
- reverse_search.universal_feature_enum_values
- reverse_search.feature_import_runs
"""

from __future__ import annotations

import json
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from core.database import supabase as sb_client

logger = logging.getLogger(__name__)

# ── Column-range → category_key mapping ────────────────────────

_COLUMN_CATEGORY_MAP: list[tuple[range, str]] = [
    (range(5, 8), "dimensions"),
    (range(8, 18), "cargo"),
    (range(18, 25), "loading"),
    (range(25, 28), "cargo"),
    (range(28, 34), "bodywork"),
    (range(34, 40), "bodywork"),
    (range(40, 45), "cabin"),
    (range(45, 51), "towing"),
    (range(51, 59), "lpg"),
    (range(59, 71), "refrigeration"),
    (range(71, 86), "lift_tachograph"),
    (range(86, 93), "lift_tachograph"),
    (range(93, 100), "drivetrain"),
    (range(100, 101), "body_type"),
    (range(101, 103), "multimedia"),
    (range(103, 109), "safety"),
    (range(109, 112), "comfort"),
    (range(112, 116), "comfort"),
    (range(116, 118), "multimedia"),
    (range(118, 120), "comfort"),
    (range(120, 121), "comfort"),
    (range(121, 132), "seats"),
    (range(132, 135), "multimedia"),
    (range(135, 140), "comfort"),
    (range(140, 143), "comfort"),
    (range(143, 146), "comfort"),
    (range(146, 148), "multimedia"),
    (range(148, 158), "comfort"),
    (range(158, 160), "safety"),
    (range(160, 162), "safety"),
    (range(162, 165), "safety"),
    (range(165, 169), "safety"),
    (range(169, 172), "security"),
    (range(172, 176), "comfort"),
    (range(176, 178), "comfort"),
    (range(178, 180), "comfort"),
]

# Columns to skip (identity, not features)
_IDENTITY_COLUMNS = {1, 2, 3, 4}

# ── Regex patterns for type detection from header ──────────────

_UNIT_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"w\s+mm", re.IGNORECASE), "mm"),
    (re.compile(r"w\s+m3", re.IGNORECASE), "m3"),
    (re.compile(r"w\s+m2", re.IGNORECASE), "m2"),
    (re.compile(r"w\s+kg", re.IGNORECASE), "kg"),
    (re.compile(r"w\s+kN", re.IGNORECASE), "kN"),
    (re.compile(r"w\s+litrach", re.IGNORECASE), "l"),
    (re.compile(r"w\s+kWh", re.IGNORECASE), "kWh"),
    (re.compile(r"w\s+km", re.IGNORECASE), "km"),
    (re.compile(r"\(kN\)", re.IGNORECASE), "kN"),
    (re.compile(r"\(kg\)", re.IGNORECASE), "kg"),
]

_BOOL_PATTERN = re.compile(r"^\s*\(tak/nie\)\s*$", re.IGNORECASE)

_ENUM_PATTERN = re.compile(
    r"^\s*\(([^)]+/[^)]+)\)\s*$",
    re.IGNORECASE,
)


@dataclass
class ParsedFeature:
    """Intermediate representation of a feature parsed from Excel."""

    column_index: int
    raw_header: str
    feature_key: str
    display_name: str
    feature_type: str  # boolean, numeric, enum, text
    canonical_unit: str | None = None
    category_key: str = "comfort"
    enum_values: list[str] = field(default_factory=list)
    description: str | None = None


def _normalize_key(text: str) -> str:
    """Convert header text to a snake_case feature_key."""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9\s]", "", text)
    text = text.strip().lower()
    text = re.sub(r"\s+", "_", text)
    return text[:80] if text else "unknown"


def _normalize_alias(text: str) -> str:
    """Normalize alias text for matching."""
    text = text.strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _get_category_for_column(col_idx: int) -> str:
    """Map 1-based column index to category_key."""
    for col_range, cat_key in _COLUMN_CATEGORY_MAP:
        if col_idx in col_range:
            return cat_key
    return "comfort"


def _detect_type_and_unit(
    header_lines: list[str],
) -> tuple[str, str | None, list[str]]:
    """Detect feature type, unit, and enum values from header text.

    Returns (feature_type, unit, enum_values).
    """
    full_text = " ".join(header_lines)

    # Check for boolean pattern in any line
    for line in header_lines[1:]:
        if _BOOL_PATTERN.match(line):
            return "boolean", None, []

    # Check for enum pattern
    for line in header_lines[1:]:
        m = _ENUM_PATTERN.match(line)
        if m:
            raw_values = m.group(1)
            values = [v.strip() for v in raw_values.split("/") if v.strip()]
            if len(values) >= 2:
                return "enum", None, values

    # Check for numeric units
    for pattern, unit in _UNIT_PATTERNS:
        if pattern.search(full_text):
            return "numeric", unit, []

    # Fallback to text
    return "text", None, []


def parse_excel_headers(file_path: str | Path) -> list[ParsedFeature]:
    """Parse Excel headers and return a list of ParsedFeature objects.

    Handles multi-line headers where the feature name is on row 1
    and the type hint (tak/nie), (enum values), or unit is on
    subsequent continuation lines.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    features: list[ParsedFeature] = []
    max_col = ws.max_column or 1

    # Read all row-1 headers
    for col_idx in range(1, max_col + 1):
        if col_idx in _IDENTITY_COLUMNS:
            continue

        raw_val = ws.cell(1, col_idx).value
        if raw_val is None:
            continue

        header_text = str(raw_val).strip()
        if not header_text:
            continue

        # Collect continuation lines (same row concept in merged cells)
        # The Excel has format hints on separate lines within same cell
        lines = [line.strip() for line in header_text.split("\n") if line.strip()]

        display_name = lines[0] if lines else header_text
        feature_type, unit, enum_values = _detect_type_and_unit(lines)
        feature_key = _normalize_key(display_name)
        category_key = _get_category_for_column(col_idx)

        # Build description from extra lines
        desc_parts = [
            part for part in lines[1:] if part and not _BOOL_PATTERN.match(part)
        ]
        description = " ".join(desc_parts) if desc_parts else None

        features.append(
            ParsedFeature(
                column_index=col_idx,
                raw_header=header_text,
                feature_key=feature_key,
                display_name=display_name,
                feature_type=feature_type,
                canonical_unit=unit,
                category_key=category_key,
                enum_values=enum_values,
                description=description,
            )
        )

    wb.close()
    logger.info("Parsed %d features from Excel", len(features))
    return features


def import_features_to_db(
    file_path: str | Path,
    imported_by: str | None = None,
) -> dict:
    """Import Excel features into Supabase reverse_search schema.

    Returns import summary dict.
    """
    features = parse_excel_headers(file_path)
    sb = sb_client

    # Fetch existing categories
    cats_resp = (
        sb.schema("reverse_search")
        .table("universal_feature_categories")
        .select("id, category_key")
        .execute()
    )
    cat_map: dict[str, str] = {r["category_key"]: r["id"] for r in cats_resp.data}

    stats = {
        "features_created": 0,
        "features_updated": 0,
        "aliases_created": 0,
        "enum_values_created": 0,
        "errors": [],
    }

    for feat in features:
        cat_id = cat_map.get(feat.category_key)
        if not cat_id:
            msg = (
                f"Category '{feat.category_key}' not found "
                f"for feature '{feat.display_name}'"
            )
            stats["errors"].append(msg)
            logger.warning(msg)
            continue

        # Upsert feature
        feature_data = {
            "feature_key": feat.feature_key,
            "display_name": feat.display_name,
            "category_id": cat_id,
            "feature_type": feat.feature_type,
            "canonical_unit": feat.canonical_unit,
            "source_standard": "fleet_excel",
            "description": feat.description,
            "sort_order": feat.column_index,
        }

        try:
            resp = (
                sb.schema("reverse_search")
                .table("universal_features")
                .upsert(feature_data, on_conflict="feature_key")
                .execute()
            )
            feature_id = resp.data[0]["id"] if resp.data else None
            if not feature_id:
                continue

            stats["features_created"] += 1

            # Create alias from raw header
            alias_data = {
                "feature_id": feature_id,
                "alias_text": feat.raw_header.split("\n")[0].strip(),
                "normalized_alias": _normalize_alias(
                    feat.raw_header.split("\n")[0],
                ),
                "source_type": "excel_import",
                "language": "pl",
            }
            try:
                sb.schema("reverse_search").table("universal_feature_aliases").upsert(
                    alias_data,
                    on_conflict="feature_id,normalized_alias",
                ).execute()
                stats["aliases_created"] += 1
            except Exception as alias_err:
                logger.warning(
                    "Alias insert error for %s: %s",
                    feat.feature_key,
                    alias_err,
                )

            # Create enum values if applicable
            for i, ev in enumerate(feat.enum_values):
                enum_data = {
                    "feature_id": feature_id,
                    "enum_key": _normalize_key(ev),
                    "display_name": ev.strip(),
                    "sort_order": (i + 1) * 10,
                }
                try:
                    sb.schema("reverse_search").table(
                        "universal_feature_enum_values"
                    ).upsert(
                        enum_data,
                        on_conflict="feature_id,enum_key",
                    ).execute()
                    stats["enum_values_created"] += 1
                except Exception as enum_err:
                    logger.warning(
                        "Enum insert error for %s/%s: %s",
                        feat.feature_key,
                        ev,
                        enum_err,
                    )

        except Exception as e:
            msg = f"Feature upsert error '{feat.feature_key}': {e}"
            stats["errors"].append(msg)
            logger.error(msg)

    # Record import run
    try:
        sb.schema("reverse_search").table("feature_import_runs").insert(
            {
                "import_type": "excel_standard",
                "file_name": Path(file_path).name,
                "source_path": str(file_path),
                "imported_by": imported_by,
                "import_summary": json.dumps(stats, ensure_ascii=False),
            }
        ).execute()
    except Exception as e:
        logger.error("Failed to record import run: %s", e)

    logger.info(
        "Import complete: %d features, %d aliases, %d enums, %d errors",
        stats["features_created"],
        stats["aliases_created"],
        stats["enum_values_created"],
        len(stats["errors"]),
    )
    return stats
