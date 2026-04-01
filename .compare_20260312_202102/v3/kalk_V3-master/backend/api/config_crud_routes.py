"""Generic config-table CRUD router: XLSX export/import + versioning.

Provides endpoints for any whitelist-approved Control Center table:
- Export to XLSX with sheet protection (headers locked, data editable)
- Import from XLSX with validation + automatic pre-import snapshot
- Version listing, manual snapshot, restore, and version export
"""

from __future__ import annotations

import io
import json
import logging
from typing import Any, List, Optional, cast

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from core.database import supabase

logger = logging.getLogger(__name__)

config_crud_router = APIRouter(tags=["Config CRUD"])

# ── Whitelist of allowed tables ──────────────────────────────────────

ALLOWED_TABLES: dict[str, str] = {
    "koszty_opon": "Koszty Opon",
    "tabela_rabaty": "Tabela Rabatów",
    "engines": "Napędy (Engines)",
    "samar_service_costs": "Koszty Serwisowe SAMAR",
    "samar_classes": "Klasy SAMAR",
    "samar_class_depreciation_rates": "Stawki Deprecjacji",
    "samar_class_mileage_corrections": "Korekty Przebiegu",
    "samar_brand_corrections": "Korekty Marki",
    "body_types": "Typy Nadwozi",
    "replacement_car_rates": "Stawki Samochodu Zastępczego",
    "service_rates_config": "Stawki Serwisowe",
    "service_base_costs_config": "Koszty Bazowe Serwisu",
    # RMS _czak tables
    "LTRAdminParametry_czak": "RMS: Parametry",
    "CennikOpon_czak": "RMS: Cennik Opon",
    "LTRAdminOplatyTransportowe_czak": "RMS: Opłaty Transportowe",
    "LTRAdminGSM_czak": "RMS: GSM",
    "TabelaSerwisowa_czak": "RMS: Tabela Serwisowa",
    "LTRAdminKosztySerwisowe_czak": "RMS: Koszty Serwisowe",
    "LTRAdminStawkaZastepczy_czak": "RMS: Stawka Zastępczego",
    "KlasaSAMAR_czak": "RMS: Klasa SAMAR",
    "LTRAdminTabelaWRKlasa_czak": "RMS: WR Klasa",
    "LTRAdminTabelaWRDeprecjacja_czak": "RMS: WR Deprecjacja",
    "LTRAdminTabelaWRPrzebieg_czak": "RMS: WR Przebieg",
    "LTRAdminTabelaWRDoposazenie_czak": "RMS: WR Doposażenie",
    "LTRAdminKorektaWRKolor_czak": "RMS: Korekta WR Kolor",
    "LTRAdminKorektaWRZabudowa_czak": "RMS: Korekta WR Zabudowa",
    "LTRAdminKorektaWRRocznik_czak": "RMS: Korekta WR Rocznik",
    "LTRAdminKorektaWRMarka_czak": "RMS: Korekta WR Marka",
    "LTRAdminUbezpieczenie_czak": "RMS: Ubezpieczenie",
    "LTRAdminWspolczynnikiSzkodowe_czak": "RMS: Współczynniki Szkodowe",
    "RodzajPaliwa_czak": "RMS: Rodzaj Paliwa",
}

# Columns that are auto-generated and should be locked in export
ID_COLUMNS = {"id", "created_at", "updated_at"}

SHEET_PASSWORD = "kalk_v3_config"

# ── Styling constants ────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(bold=True, size=11)
_ID_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
_LOCKED = Protection(locked=True)
_UNLOCKED = Protection(locked=False)


# ── Pydantic models ─────────────────────────────────────────────────


class VersionInfo(BaseModel):
    id: int
    table_name: str
    version_num: int
    label: str
    created_at: str
    created_by: str
    row_count: int = 0


class SnapshotRequest(BaseModel):
    label: str = ""


class ImportReport(BaseModel):
    status: str
    updated: int = 0
    inserted: int = 0
    skipped: int = 0
    errors: List[str] = []
    snapshot_version: Optional[int] = None


# ── Helpers ──────────────────────────────────────────────────────────


def _validate_table(table_name: str) -> str:
    """Validate table name against whitelist, return label."""
    if table_name not in ALLOWED_TABLES:
        raise HTTPException(
            status_code=400,
            detail=f"Tabela '{table_name}' nie jest dozwolona. "
            f"Dostępne: {', '.join(sorted(ALLOWED_TABLES.keys()))}",
        )
    return ALLOWED_TABLES[table_name]


def _fetch_all_rows(table_name: str) -> list[dict[str, Any]]:
    """Fetch all rows from a table via Supabase."""
    response = supabase.table(table_name).select("*").execute()
    return cast(Any, response.data) or []


def _get_next_version_num(table_name: str) -> int:
    """Get the next version number for a table."""
    response = (
        supabase.table("config_table_versions")
        .select("version_num")
        .eq("table_name", table_name)
        .order("version_num", desc=True)
        .limit(1)
        .execute()
    )
    data = cast(Any, response.data)
    if data:
        return int(data[0]["version_num"]) + 1
    return 1


def _create_snapshot(
    table_name: str, label: str = "", created_by: str = "system"
) -> int:
    """Create a snapshot of the current table data. Returns version_num."""
    rows = _fetch_all_rows(table_name)
    version_num = _get_next_version_num(table_name)

    supabase.table("config_table_versions").insert(
        {
            "table_name": table_name,
            "version_num": version_num,
            "label": label,
            "snapshot": json.dumps(rows, default=str),
            "created_by": created_by,
        }
    ).execute()

    logger.info(
        "Snapshot v%d created for %s (%d rows)",
        version_num,
        table_name,
        len(rows),
    )
    return version_num


def _build_protected_xlsx(rows: list[dict[str, Any]], sheet_title: str) -> io.BytesIO:
    """Build an XLSX workbook with sheet protection.

    - Header row: locked, grey background, bold
    - ID/auto-generated columns: locked, yellow background
    - Data cells: UNLOCKED (editable)
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = sheet_title[:31]  # Excel 31-char limit

    if not rows:
        ws.append(["Brak danych"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    headers = list(rows[0].keys())

    # Identify which columns are ID/auto columns
    id_col_indices: set[int] = set()
    for idx, header in enumerate(headers):
        if header.lower() in ID_COLUMNS:
            id_col_indices.add(idx)

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.protection = _LOCKED
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if (col_idx - 1) in id_col_indices:
                cell.fill = _ID_FILL
                cell.protection = _LOCKED
            else:
                cell.protection = _UNLOCKED

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(rows) + 2, 52)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    # Enable sheet protection (only unlocked cells are editable)
    ws.protection.sheet = True
    ws.protection.password = SHEET_PASSWORD
    ws.protection.enable()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── XLSX Export ───────────────────────────────────────────────────────


@config_crud_router.get("/config/{table_name}/export-xlsx")
async def export_table_xlsx(table_name: str) -> StreamingResponse:
    """Export a config table as a protected XLSX file."""
    label = _validate_table(table_name)

    try:
        rows = _fetch_all_rows(table_name)
        output = _build_protected_xlsx(rows, label)

        safe_name = table_name.replace(" ", "_")
        filename = f"{safe_name}_export.xlsx"
        return StreamingResponse(
            output,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
    except Exception as exc:
        logger.exception("Export failed for %s", table_name)
        raise HTTPException(status_code=500, detail=str(exc))


# ── XLSX Import ───────────────────────────────────────────────────────


@config_crud_router.post("/config/{table_name}/import-xlsx")
async def import_table_xlsx(
    table_name: str, file: UploadFile = File(...)
) -> ImportReport:
    """Import XLSX into a config table.

    1. Validates column headers match current schema
    2. Creates automatic pre-import snapshot
    3. Upserts rows: ID present → UPDATE, ID absent → INSERT
    """
    _validate_table(table_name)

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Plik XLSX nie zawiera arkuszy")

        # Read headers from first row
        file_headers: list[str] = []
        for cell in ws[1]:
            if cell.value is not None:
                file_headers.append(str(cell.value))

        if not file_headers:
            raise HTTPException(status_code=400, detail="Brak nagłówków w pliku XLSX")

        # Validate against current DB schema
        current_rows = _fetch_all_rows(table_name)
        if current_rows:
            db_headers = set(current_rows[0].keys())
            unknown = set(file_headers) - db_headers
            if unknown:
                raise HTTPException(
                    status_code=400,
                    detail=f"Nieznane kolumny w pliku: {', '.join(unknown)}",
                )

        # Create pre-import snapshot
        snapshot_version = _create_snapshot(
            table_name, label="Auto: przed importem XLSX"
        )

        # Process rows
        updated = 0
        inserted = 0
        skipped = 0
        errors: list[str] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data: dict[str, Any] = {}
            is_empty = True

            for col_idx, cell in enumerate(row):
                if col_idx >= len(file_headers):
                    break
                header = file_headers[col_idx]
                value = cell.value
                if value is not None:
                    is_empty = False
                row_data[header] = value

            if is_empty:
                skipped += 1
                continue

            try:
                row_id = row_data.get("id")
                if row_id is not None and str(row_id).strip() != "":
                    # UPDATE existing row
                    update_data = {
                        k: v for k, v in row_data.items() if k.lower() not in ID_COLUMNS
                    }
                    supabase.table(table_name).update(update_data).eq(
                        "id", row_id
                    ).execute()
                    updated += 1
                else:
                    # INSERT new row (remove ID so DB auto-generates)
                    insert_data = {
                        k: v for k, v in row_data.items() if k.lower() not in ID_COLUMNS
                    }
                    supabase.table(table_name).insert(insert_data).execute()
                    inserted += 1
            except Exception as row_exc:
                errors.append(f"Wiersz {row_idx}: {row_exc}")

        return ImportReport(
            status="success",
            updated=updated,
            inserted=inserted,
            skipped=skipped,
            errors=errors,
            snapshot_version=snapshot_version,
        )

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Import failed for %s", table_name)
        raise HTTPException(status_code=500, detail=str(exc))


# ── Versioning ────────────────────────────────────────────────────────


@config_crud_router.get("/config/{table_name}/versions")
async def list_versions(table_name: str) -> list[VersionInfo]:
    """List all saved versions/snapshots for a config table."""
    _validate_table(table_name)

    try:
        response = (
            supabase.table("config_table_versions")
            .select(
                "id, table_name, version_num, label, created_at, created_by, snapshot"
            )
            .eq("table_name", table_name)
            .order("version_num", desc=True)
            .execute()
        )
        data = cast(Any, response.data) or []

        results: list[VersionInfo] = []
        for row in data:
            snapshot = row.get("snapshot")
            row_count = 0
            if isinstance(snapshot, str):
                try:
                    row_count = len(json.loads(snapshot))
                except Exception:
                    pass
            elif isinstance(snapshot, list):
                row_count = len(snapshot)

            results.append(
                VersionInfo(
                    id=row["id"],
                    table_name=row["table_name"],
                    version_num=row["version_num"],
                    label=row.get("label", ""),
                    created_at=str(row["created_at"]),
                    created_by=row.get("created_by", "system"),
                    row_count=row_count,
                )
            )
        return results
    except Exception as exc:
        logger.exception("Failed to list versions for %s", table_name)
        raise HTTPException(status_code=500, detail=str(exc))


@config_crud_router.post("/config/{table_name}/snapshot")
async def create_manual_snapshot(
    table_name: str, body: SnapshotRequest | None = None
) -> dict[str, Any]:
    """Manually create a named snapshot of current table state."""
    _validate_table(table_name)
    label = body.label if body else ""

    try:
        version_num = _create_snapshot(table_name, label=label or "Ręczny snapshot")
        return {
            "status": "success",
            "version_num": version_num,
            "table_name": table_name,
        }
    except Exception as exc:
        logger.exception("Snapshot creation failed for %s", table_name)
        raise HTTPException(status_code=500, detail=str(exc))


@config_crud_router.post("/config/{table_name}/restore/{version_id}")
async def restore_version(table_name: str, version_id: int) -> dict[str, Any]:
    """Restore table data from a specific version snapshot.

    1. Creates a 'before restore' snapshot of current data
    2. Deletes all current rows
    3. Inserts rows from the snapshot
    """
    _validate_table(table_name)

    try:
        # Fetch the target snapshot
        response = (
            supabase.table("config_table_versions")
            .select("snapshot, version_num")
            .eq("id", version_id)
            .eq("table_name", table_name)
            .single()
            .execute()
        )
        version_data = cast(Any, response.data)
        if not version_data:
            raise HTTPException(
                status_code=404,
                detail=f"Wersja {version_id} nie istnieje dla {table_name}",
            )

        snapshot_raw = version_data["snapshot"]
        if isinstance(snapshot_raw, str):
            snapshot_rows = json.loads(snapshot_raw)
        else:
            snapshot_rows = snapshot_raw

        source_version = version_data["version_num"]

        # Create safety snapshot before restore
        _create_snapshot(
            table_name,
            label=f"Auto: przed przywróceniem v{source_version}",
        )

        # Delete all current rows
        # Supabase requires a filter for delete, use gte on id
        supabase.table(table_name).delete().gte("id", 0).execute()

        # Re-insert snapshot rows
        restored_count = 0
        if snapshot_rows:
            # Insert in batches of 100
            batch_size = 100
            for i in range(0, len(snapshot_rows), batch_size):
                batch = snapshot_rows[i : i + batch_size]
                supabase.table(table_name).insert(batch).execute()
                restored_count += len(batch)

        return {
            "status": "success",
            "restored_version": source_version,
            "restored_rows": restored_count,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Restore failed for %s v%d", table_name, version_id)
        raise HTTPException(status_code=500, detail=str(exc))


@config_crud_router.get("/config/{table_name}/versions/{version_id}/export-xlsx")
async def export_version_xlsx(table_name: str, version_id: int) -> StreamingResponse:
    """Export a specific historical version as XLSX."""
    label = _validate_table(table_name)

    try:
        response = (
            supabase.table("config_table_versions")
            .select("snapshot, version_num")
            .eq("id", version_id)
            .eq("table_name", table_name)
            .single()
            .execute()
        )
        version_data = cast(Any, response.data)
        if not version_data:
            raise HTTPException(
                status_code=404,
                detail=f"Wersja {version_id} nie istnieje",
            )

        snapshot_raw = version_data["snapshot"]
        if isinstance(snapshot_raw, str):
            rows = json.loads(snapshot_raw)
        else:
            rows = snapshot_raw

        version_num = version_data["version_num"]
        sheet_title = f"{label} v{version_num}"
        output = _build_protected_xlsx(rows, sheet_title)

        safe_name = table_name.replace(" ", "_")
        filename = f"{safe_name}_v{version_num}.xlsx"
        return StreamingResponse(
            output,
            headers={"Content-Disposition": f"attachment; filename={filename}"},
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Version export failed")
        raise HTTPException(status_code=500, detail=str(exc))
