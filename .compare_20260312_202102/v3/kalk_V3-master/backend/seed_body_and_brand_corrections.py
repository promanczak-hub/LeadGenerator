"""
Seed script: generuje ~16k wierszy body_type_wr_corrections
(klasa × marka × body_type) + importuje brand corrections z Excela.

Uruchomienie:
    python seed_body_and_brand_corrections.py
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from core.database import supabase  # noqa: E402

from seed_depreciation_rates import (  # noqa: E402
    CLASS_MAP,
    FUEL_SUFFIX_MAP,
)

EXCEL_PATH = Path(
    r"C:\Users\proma\Downloads"
    r"\DRAFT KALKULATORA WARTOŚCI REZYDUALNYCH"
    r" ver aktualna JŁ 02.02 (version 1).xlsx"
)


def fetch_all_brands() -> list[str]:
    """Pobiera unikalne marki z vehicle_synthesis."""
    try:
        res = supabase.table("vehicle_synthesis").select("brand").execute()
        brands: set[str] = set()
        for row in res.data or []:
            b = (row.get("brand") or "").strip().upper()
            if b:
                brands.add(b)
        # Dodaj marki z Excela KOR. MARKA
        wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
        ws = wb["KOR. MARKA"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            brand = row[1]
            if brand and isinstance(brand, str):
                brands.add(brand.strip().upper())
        return sorted(brands)
    except Exception as e:
        print(f"⚠ Błąd pobierania marek: {e}")
        return []


def fetch_samar_classes() -> list[dict[str, Any]]:
    """Pobiera wszystkie klasy SAMAR."""
    res = supabase.table("samar_classes").select("id, name").execute()
    return res.data or []


def fetch_body_types() -> list[dict[str, Any]]:
    """Pobiera wszystkie typy nadwozi."""
    res = supabase.table("body_types").select("id, name").execute()
    return res.data or []


def seed_body_type_corrections() -> None:
    """Generuje wpisy body_type_wr_corrections (default 0)."""
    classes = fetch_samar_classes()
    body_types = fetch_body_types()
    brands = fetch_all_brands()

    if not brands:
        brands = [""]  # Przynajmniej jeden wpis bez marki

    total = len(classes) * len(brands) * len(body_types)
    print(f"\n📊 Generowanie {total} wierszy body_type_wr_corrections...")
    print(
        f"   Klasy: {len(classes)}, Marki: {len(brands)}, Nadwozia: {len(body_types)}"
    )

    batch_size = 500
    records: list[dict[str, Any]] = []
    inserted = 0

    for cls in classes:
        for brand in brands:
            for bt in body_types:
                records.append(
                    {
                        "samar_class_id": cls["id"],
                        "brand_name": brand,
                        "body_type_id": bt["id"],
                        "correction_percent": 0,
                        "zabudowa_correction_percent": 0,
                    }
                )

                if len(records) >= batch_size:
                    _upsert_batch(records)
                    inserted += len(records)
                    print(f"   ✅ {inserted}/{total}")
                    records = []

    if records:
        _upsert_batch(records)
        inserted += len(records)
        print(f"   ✅ {inserted}/{total}")

    print(f"   🎉 Zakończono: {inserted} wierszy")


def _upsert_batch(records: list[dict[str, Any]]) -> None:
    """Upsert batch do body_type_wr_corrections."""
    supabase.table("body_type_wr_corrections").upsert(
        records,
        on_conflict="samar_class_id,brand_name,body_type_id",
    ).execute()


def seed_brand_corrections_from_excel() -> None:
    """Importuje korekty marek z Excela KOR. MARKA do ltr_admin_korekta_wr_markas."""
    print("\n📖 Importowanie korekt marek z Excela...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb["KOR. MARKA"]

    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        class_code = row[0]
        brand = row[1]
        fuel = row[2]
        korekta = row[4]

        if not class_code or not brand:
            continue
        if not isinstance(class_code, str):
            continue

        samar_id = _resolve_class(str(class_code))
        if samar_id is None:
            print(f"   ⚠ Pominięto: klasa '{class_code}' nieznana")
            continue

        fuel_id = FUEL_SUFFIX_MAP.get(str(fuel))
        if fuel_id is None:
            print(f"   ⚠ Pominięto: silnik '{fuel}' nieznany")
            continue

        records.append(
            {
                "klasa_wr_id": samar_id,
                "rodzaj_paliwa": fuel_id,
                "brand_name": str(brand).strip().upper(),
                "korekta_procent": float(korekta) if korekta else 0.0,
            }
        )

    print(f"   Wczytano {len(records)} korekt marek")

    if records:
        # Upsert in batches
        batch_size = 100
        for i in range(0, len(records), batch_size):
            batch = records[i : i + batch_size]
            supabase.table("ltr_admin_korekta_wr_markas").upsert(
                batch,
                on_conflict="klasa_wr_id,rodzaj_paliwa,brand_name",
            ).execute()
        print(f"   ✅ Załadowano {len(records)} korekt")


def _resolve_class(code: str) -> int | None:
    """Rozwiązuje kod klasy Excela na samar_class_id."""
    result = CLASS_MAP.get(code)
    if result:
        return result
    # Case-insensitive fallback
    for k, v in CLASS_MAP.items():
        if k.upper() == code.upper():
            return v
    return None


def main() -> None:
    print("🚀 Seed: Body Type Corrections + Brand Corrections")

    # 1. Brand corrections from Excel
    seed_brand_corrections_from_excel()

    # 2. Body type corrections (16k rows)
    seed_body_type_corrections()

    print("\n🎉 Seed zakończony!")


if __name__ == "__main__":
    main()
