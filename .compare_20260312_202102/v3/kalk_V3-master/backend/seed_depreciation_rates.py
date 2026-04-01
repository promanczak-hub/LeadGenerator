"""
Import danych deprecjacji z Excela DRAFT KALKULATORA WARTOŚCI REZYDUALNYCH
do tabel Supabase.

Pipeline obliczania RV (z Excela):
  1. 48WR% = TAB.WR KLASA[klasa+silnik]        → bazowa retencja po 48mies/140k
  2. Korekta Marki = KOR. MARKA[klasa+silnik+marka]
  3. WR_baza = cena_kat × (48WR% + korekta_marki)
  4. Korekta Okresu = TAB. OKRES FINAL[klasa+silnik][rok]
     → rok < 4: WR rośnie (aprecjacja)
     → rok = 4: 0 (punkt bazowy)
     → rok > 4: WR maleje (deprecjacja)
  5. Doposażenie = TAB. DOPOSAŻENIA[klasa+silnik][rok] × cena_opcji
  6. Korekta Przebiegu = TAB. PRZEBIEG[klasa]
     → odchylenie przebieg vs 140k, z progiem 190k
  7. Korekty: KOLOR, NADWOZIE, ROCZNIK
  8. Finał: WR = WR_baza × (1 ± okres) + opcje ± przebieg ± korekty
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl
import requests

from core.database import supabase

EXCEL_PATH = Path(
    r"C:\Users\proma\Downloads"
    r"\DRAFT KALKULATORA WARTOŚCI REZYDUALNYCH"
    r" ver aktualna JŁ 02.02 (version 1).xlsx"
)
BASE_URL = "http://127.0.0.1:8000"

# ── Excel fuel suffix → Supabase engine_id ─────────────────────────
FUEL_SUFFIX_MAP: dict[str, int] = {
    "Pb": 1,  # Benzyna
    "PB": 1,
    "ON": 2,  # Diesel
    "EV": 7,  # Elektryczny (BEV)
    "PHEV": 6,  # Plug-in Hybrid
    "HEV": 5,  # Hybryda
}

# Engine IDs that inherit from others (brak w Excelu)
INHERIT_MAP: dict[int, int] = {
    3: 1,  # Benzyna mHEV ← Benzyna
    4: 2,  # Diesel mHEV  ← Diesel
    8: 7,  # FCEV         ← BEV
    9: 1,  # LPG          ← Benzyna
}

ALL_ENGINE_IDS = [1, 2, 3, 4, 5, 6, 7, 8, 9]

# ── Excel class code → Supabase samar_class_id ─────────────────────
CLASS_MAP: dict[str, int] = {
    "A": 1,
    "Asport": 12,
    "B": 2,
    "Bsport": 13,
    "Bvan": 15,
    "Bsuv": 14,
    "C": 3,
    "Csport": 16,
    "Cvan": 18,
    "Csuv": 17,
    "D": 4,
    "Dsport": 19,
    "Dvan": 21,
    "Dsuv": 20,
    "E": 5,
    "Esuv": 22,
    "ESUV": 22,  # Excel uppercase variant
    "F": 6,
    "Fsport": 23,
    "Fsuv": 24,
    "M": 25,
    "Mvan": 26,
    "R": 27,
    "P": 28,
    "T PICK-UP": 29,
}


def _parse_class_fuel(cell_value: str) -> tuple[str, int] | None:
    """Rozbija np. 'CsuvPb' → ('Csuv', 1) lub 'T PICK-UPON' → ('T PICK-UP', 2)."""
    if not cell_value or not isinstance(cell_value, str):
        return None

    # Specjalna obsługa T PICK-UP
    if cell_value.startswith("T PICK-UP"):
        suffix = cell_value[len("T PICK-UP") :]
        eid = FUEL_SUFFIX_MAP.get(suffix)
        if eid is not None:
            return ("T PICK-UP", eid)
        return None

    # Próbuj suffiksy od najdłuższego
    for suffix in ("PHEV", "HEV", "EV", "ON", "Pb", "PB"):
        if cell_value.endswith(suffix):
            class_code = cell_value[: -len(suffix)]
            eid = FUEL_SUFFIX_MAP[suffix]
            return (class_code, eid)
    return None


def read_wr_klasa(
    wb: openpyxl.Workbook,
) -> dict[tuple[int, int], float]:
    """TAB.WR KLASA → {(samar_class_id, engine_id): rv_percent}."""
    ws = wb["TAB.WR KLASA"]
    result: dict[tuple[int, int], float] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_val = row[0]
        rv_pct = row[1]
        if not cell_val or not isinstance(cell_val, str):
            continue
        # Skip fuel headers
        if cell_val in ("BENZYNA", "DIESEL", "EV", "PHEV", "HEV"):
            continue

        parsed = _parse_class_fuel(cell_val)
        if parsed is None:
            print(f"  ⚠ WR KLASA: nie rozpoznano '{cell_val}'")
            continue

        class_code, engine_id = parsed
        samar_id = CLASS_MAP.get(class_code)
        if samar_id is None:
            print(f"  ⚠ WR KLASA: brak mapowania klasy '{class_code}'")
            continue

        result[(samar_id, engine_id)] = float(rv_pct) if rv_pct else 0.0

    return result


def read_okres_final(
    wb: openpyxl.Workbook,
) -> dict[tuple[int, int], list[float]]:
    """TAB. OKRES FINAL → {(samar_id, engine_id): [rok0..rok7]}."""
    ws = wb["TAB. OKRES FINAL"]
    result: dict[tuple[int, int], list[float]] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        cell_val = row[0]
        if not cell_val or not isinstance(cell_val, str):
            continue
        if cell_val in ("BENZYNA", "DIESEL", "EV", "PHEV", "HEV"):
            continue

        parsed = _parse_class_fuel(cell_val)
        if parsed is None:
            continue

        class_code, engine_id = parsed
        samar_id = CLASS_MAP.get(class_code)
        if samar_id is None:
            continue

        # Kolumny B-I = rok 0-7
        rates = [float(row[i]) if row[i] is not None else 0.0 for i in range(1, 9)]
        result[(samar_id, engine_id)] = rates

    return result


def read_doposazenia(
    wb: openpyxl.Workbook,
) -> dict[tuple[int, int], list[float]]:
    """TAB. DOPOSAŻENIA → {(samar_id, engine_id): [rok0..rok7]}."""
    ws = wb["TAB. DOPOSAŻENIA"]
    result: dict[tuple[int, int], list[float]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_val = row[0]
        if not cell_val or not isinstance(cell_val, str):
            continue
        if cell_val in ("KLASY",):
            continue

        parsed = _parse_class_fuel(cell_val)
        if parsed is None:
            continue

        class_code, engine_id = parsed
        samar_id = CLASS_MAP.get(class_code)
        if samar_id is None:
            continue

        rates = [float(row[i]) if row[i] is not None else 0.0 for i in range(1, 9)]
        result[(samar_id, engine_id)] = rates

    return result


def read_przebieg(
    wb: openpyxl.Workbook,
) -> dict[int, tuple[float, float]]:
    """TAB. PRZEBIEG → {samar_id: (under_190, over_190)}."""
    ws = wb["TAB. PRZEBIEG"]
    result: dict[int, tuple[float, float]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        class_code = row[0]
        if not class_code or not isinstance(class_code, str):
            continue

        samar_id = CLASS_MAP.get(class_code)
        if samar_id is None:
            print(f"  ⚠ PRZEBIEG: brak mapowania '{class_code}'")
            continue

        under = float(row[1]) if row[1] else 0.0
        over = float(row[2]) if row[2] else 0.0
        result[samar_id] = (under, over)

    return result


def read_marka(
    wb: openpyxl.Workbook,
) -> list[dict[str, Any]]:
    """KOR. MARKA → lista rekordów brand correction."""
    ws = wb["KOR. MARKA"]
    result: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        class_code = row[0]
        brand = row[1]
        fuel = row[2]
        korekta = row[4]  # Kolumna 5 (E)

        if not class_code or not brand:
            continue
        if not isinstance(class_code, str):
            continue

        samar_id = CLASS_MAP.get(class_code.upper(), CLASS_MAP.get(class_code))
        if samar_id is None:
            # Próbuj case-insensitive
            for k, v in CLASS_MAP.items():
                if k.upper() == class_code.upper():
                    samar_id = v
                    break
        if samar_id is None:
            continue

        fuel_id = FUEL_SUFFIX_MAP.get(fuel)
        if fuel_id is None:
            continue

        result.append(
            {
                "klasa_wr_id": samar_id,
                "rodzaj_paliwa": fuel_id,
                "marka_name": str(brand),
                "korekta_procent": float(korekta) if korekta else 0.0,
            }
        )

    return result


def build_depreciation_records(
    wr_klasa: dict[tuple[int, int], float],
    okres: dict[tuple[int, int], list[float]],
    doposazenia: dict[tuple[int, int], list[float]],
) -> list[dict[str, Any]]:
    """Buduje rekordy do samar_class_depreciation_rates."""
    records: list[dict[str, Any]] = []

    # Zbierz wszystkie unikalne (samar_id, engine_id)
    all_keys: set[tuple[int, int]] = set()
    all_keys.update(wr_klasa.keys())
    all_keys.update(okres.keys())

    for samar_id, engine_id in sorted(all_keys):
        base_rv = wr_klasa.get((samar_id, engine_id), 0.0)
        period_rates = okres.get(
            (samar_id, engine_id),
            [0.0] * 8,
        )
        opt_rates = doposazenia.get(
            (samar_id, engine_id),
            [0.0] * 8,
        )

        # Rok 0 = bazowe RV% (z TAB.WR KLASA)
        records.append(
            {
                "samar_class_id": samar_id,
                "fuel_type_id": engine_id,
                "year": 0,
                "base_depreciation_percent": round(base_rv, 4),
                "options_depreciation_percent": round(
                    opt_rates[0] if len(opt_rates) > 0 else 0.0,
                    4,
                ),
            }
        )

        # Rok 1-7 = korekty okresowe
        for yr in range(1, 8):
            period_val = period_rates[yr] if yr < len(period_rates) else 0.0
            opt_val = opt_rates[yr] if yr < len(opt_rates) else 0.0

            records.append(
                {
                    "samar_class_id": samar_id,
                    "fuel_type_id": engine_id,
                    "year": yr,
                    "base_depreciation_percent": round(period_val, 4),
                    "options_depreciation_percent": round(opt_val, 4),
                }
            )

    return records


def inherit_missing_engines(
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Kopiuje dane dla silników mHEV/FCEV/LPG z najbliższych odpowiedników."""
    # Index existing records
    existing: dict[tuple[int, int, int], dict] = {}
    for r in records:
        key = (r["samar_class_id"], r["fuel_type_id"], r["year"])
        existing[key] = r

    # Zbierz unikalne samar_class_ids
    class_ids = {r["samar_class_id"] for r in records}

    extras: list[dict[str, Any]] = []
    for samar_id in class_ids:
        for target_eid, source_eid in INHERIT_MAP.items():
            for yr in range(8):
                target_key = (samar_id, target_eid, yr)
                if target_key in existing:
                    continue  # Już istnieje

                source_key = (samar_id, source_eid, yr)
                source = existing.get(source_key)
                if source is None:
                    continue

                extras.append(
                    {
                        "samar_class_id": samar_id,
                        "fuel_type_id": target_eid,
                        "year": yr,
                        "base_depreciation_percent": source[
                            "base_depreciation_percent"
                        ],
                        "options_depreciation_percent": source[
                            "options_depreciation_percent"
                        ],
                    }
                )

    return records + extras


def build_mileage_records(
    przebieg: dict[int, tuple[float, float]],
) -> list[dict[str, Any]]:
    """Buduje rekordy do samar_class_mileage_corrections."""
    records: list[dict[str, Any]] = []
    for samar_id, (under, over) in przebieg.items():
        for eid in ALL_ENGINE_IDS:
            records.append(
                {
                    "samar_class_id": samar_id,
                    "fuel_type_id": eid,
                    "under_threshold_percent": round(under, 5),
                    "over_threshold_percent": round(over, 5),
                }
            )
    return records


def api_post(
    endpoint: str,
    data: list[dict],
    batch_size: int = 100,
) -> None:
    """Wysyła dane w batchach do API."""
    total = len(data)
    sent = 0
    for i in range(0, total, batch_size):
        batch = data[i : i + batch_size]
        resp = requests.post(
            f"{BASE_URL}{endpoint}",
            json=batch,
            timeout=30,
        )
        if resp.status_code >= 400:
            print(f"  ❌ Błąd batch {i}: {resp.text[:200]}")
            sys.exit(1)
        sent += len(batch)
        print(f"  ✅ {sent}/{total}")


def main() -> None:
    print(f"📖 Wczytuję Excel: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # 1. TAB.WR KLASA
    print("\n1️⃣  TAB.WR KLASA (bazowe RV%)...")
    wr_klasa = read_wr_klasa(wb)
    print(f"   Wczytano {len(wr_klasa)} pozycji")

    # 2. TAB. OKRES FINAL
    print("\n2️⃣  TAB. OKRES FINAL (korekty okresowe)...")
    okres = read_okres_final(wb)
    print(f"   Wczytano {len(okres)} pozycji")

    # 3. TAB. DOPOSAŻENIA
    print("\n3️⃣  TAB. DOPOSAŻENIA (deprecjacja opcji)...")
    doposazenia = read_doposazenia(wb)
    print(f"   Wczytano {len(doposazenia)} pozycji")

    # 4. TAB. PRZEBIEG
    print("\n4️⃣  TAB. PRZEBIEG (korekty przebiegu)...")
    przebieg = read_przebieg(wb)
    print(f"   Wczytano {len(przebieg)} klas")

    # 5. KOR. MARKA
    print("\n5️⃣  KOR. MARKA (korekty marek)...")
    marka = read_marka(wb)
    print(f"   Wczytano {len(marka)} korekt marek")

    # ── Budowanie rekordów ──
    print("\n📊 Budowanie rekordów deprecjacji...")
    depr_records = build_depreciation_records(wr_klasa, okres, doposazenia)
    print(f"   Bazowych: {len(depr_records)}")

    depr_records = inherit_missing_engines(depr_records)
    print(f"   Po dziedziczeniu (mHEV/FCEV/LPG): {len(depr_records)}")

    print("\n📊 Budowanie rekordów przebiegu...")
    mileage_records = build_mileage_records(przebieg)
    print(f"   Rekordów: {len(mileage_records)}")

    # ── Czyszczenie starych danych ──
    print("\n🗑️  Czyszczenie starych danych deprecjacji...")
    supabase.table("samar_class_depreciation_rates").delete().neq(
        "id",
        0,
    ).execute()
    print("   ✅ Wyczyszczono")

    # ── Wysyłanie do API ──
    print("\n🚀 Wysyłanie deprecjacji do API...")
    api_post("/api/depreciation-rates/bulk", depr_records)

    print("\n🚀 Wysyłanie korekt przebiegu do Supabase...")
    # Direct upsert (no bulk API for mileage corrections)
    batch_size = 100
    total_m = len(mileage_records)
    sent_m = 0
    for i in range(0, total_m, batch_size):
        batch = mileage_records[i : i + batch_size]
        supabase.table("samar_class_mileage_corrections").upsert(
            batch,
            on_conflict="samar_class_id,fuel_type_id",
        ).execute()
        sent_m += len(batch)
        print(f"  ✅ {sent_m}/{total_m}")

    # Brand corrections - log only (need marka_id mapping)
    print(f"\n📋 Korekty marek: {len(marka)} (do ręcznej obsługi)")
    for m in marka[:5]:
        print(f"   {m}")
    if len(marka) > 5:
        print(f"   ... i {len(marka) - 5} więcej")

    print("\n🎉 Import zakończony!")


if __name__ == "__main__":
    main()
